import datetime
import os
import sys
import re
import string
import xlsxwriter
import openpyxl
from time import sleep
from netmiko import ConnectHandler, NetmikoTimeoutException, NetmikoAuthenticationException



password = 'Airtel@1q2w' # Tanzania
#password = 'A1rt3l@123' # Uganda 
username = 'chugorji'
passwordfw = 'Bharti@1289'
usernamefw = 'admin'

filetime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
folder = '/SWAP4/'
home = os.path.dirname(sys.argv[0]).replace('\\', '/')
fpath = home + folder
if not os.path.exists(fpath):
    os.mkdir(fpath)
# mynodes = open(home + "/myfiles/ugnodes").readlines()
# myconfig = open(home + "/myfiles/ugconfig").readlines()
mynodes = open(home + "/myfiles/mynodes").readlines() # Tanzania
#mynodes = open(home + "/myfiles/Allnodes").readlines() # Tanzania
myconfig = open(home + "/myfiles/myconfig").readlines() # Tanzania

workbook = xlsxwriter.Workbook(fpath+'MyReport'+filetime+'.xlsx')
worksheet = workbook.add_worksheet('Result')
workbook.close()
wb = openpyxl.load_workbook(fpath+'MyReport'+filetime+'.xlsx')
sheet = wb['Result']


print(mynodes)


def excellog(row, col, log):
    printable = set(string.printable)
    filterlog = ''.join(filter(lambda x: x in printable, log))

    sheet.cell(row=row, column=col).value = filterlog
    wb.save(fpath+'MyReport'+filetime+'.xlsx')
    wb.close()


oneoff = open(fpath + 'oneoff_' + filetime, 'w+')

row = 0
for each in mynodes:
    print(each)
    row += 1
    ip = re.split('\t', each)
    if each.strip() != "" or each.strip() != None:
        print(ip[0])
        #file = open(fpath+ip[0], 'w+')

        cisco = {
            'host': ip[1].strip('\n') ,
            'device_type': 'cisco_ios',
            'username': username,
            'password': password,
            'session_log' : f'{fpath+ip[0]}_{filetime}.log'
        }

    try:
        con = ConnectHandler(**cisco)
        con.enable()
        col=1
        sheet.max_row
        sheet.cell(row=row, column=col).value = ip[0] 
        for eachConfig in myconfig:
            
            res = con.send_command(eachConfig, read_timeout=180.0, expect_string=r'#')
            print(con.find_prompt() + eachConfig, '\n', res, '\n\n')
            col+=1  
            sleep(2)   
            eres = con.find_prompt() + eachConfig +'\n' + res +'\n\n'
            excellog(row, col, eres)
            sleep(2)   
            #file.write(con.find_prompt() + '\n' + res +'\n\n')
            #sleep(2)
            oneoff.write(con.find_prompt() + eachConfig +'\n' + res +'\n\n')
            sleep(2)
        oneoff.write('\n\n===============\n' +
                         ' '.join(ip) + '\n===============\n')
        
        con.disconnect()

    except (NetmikoTimeoutException, NetmikoAuthenticationException):
        col = 1
        sheet.max_row
        sheet.cell(row=row, column=col).value = re.split('\t', each)[0]
        sheet.cell(row=row, column=2).value = NetmikoTimeoutException
        print(re.split('\t', each)[0] + 'SSH Connection failed')
        file.write('SSH Connection failed' + '\n'+ NetmikoTimeoutException)
        continue
