import datetime
import socket
import telnetlib
import os
import sys
import re
import paramiko
import string
import xlsxwriter
import openpyxl
from time import sleep
from paramiko import SSHException, AuthenticationException

password = 'Airtel@123'
username = 'chugorji'
filetime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
activity = 'precheck'
# activity = 'postcheck'
folder = '/Manoj_SRTE/'

home = os.path.dirname(sys.argv[0]).replace('\\', '/')
print(home)
mynodes = open(home + r"/myfiles/mynodes").readlines()
#mynodes = open(home + "/myfiles/allnodes").readlines()
bukconfig = open(home + r"/myfiles/"+activity).read()
fpath = home + folder
fType = fpath + activity
print(fType)

if not os.path.exists(fpath):
    os.makedirs(fType)

workbook = xlsxwriter.Workbook(
    fType+'/myReport_'+filetime+'.xlsx')
worksheet = workbook.add_worksheet('Result')
workbook.close()
wb = openpyxl.load_workbook(
    fType+'/myReport_'+filetime+'.xlsx')
sheet = wb['Result']
mylog = open(fType + '/myReport_'+filetime+'.txt', 'w+')

print(mynodes)


def excellog(row, col, log):
    printable = set(string.printable)
    filterlog = ''.join(filter(lambda x: x in printable, log))

    sheet.cell(row=row, column=col).value = filterlog
    wb.save(fType+'/myReport_'+filetime+'.xlsx')
    wb.close()


row = 0
for each in mynodes:
    print(each)
    row += 1
    ip = re.split('\t', each)
    if ip != "\n" or ip != None:
        file = open(fpath+ip[0]+'_'+filetime, 'w+')
        cisco = {
            'hostname': ip[1].strip('\n'),
            'port': 22,
            'username': username,
            'password': password,
            'timeout': 30
        }
    connected = True

    try:
        con = paramiko.SSHClient()
        con.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        # print(cisco)
        con.connect(**cisco)
        waittime = 0
        proceed = True
        with con.invoke_shell() as shell:
            shell.recv(10000)
            print('Wait for connection', '')
            while True:
                if shell.recv_ready():
                    sleep(1)
                    print('')
                    shell.send('terminal len 0\r')
                    print(shell.recv(10000))
                    break
                else :
                    waittime += 1
                    print('.', end ="")
                    sleep(1)
                    if waittime >= 25:
                        con.close()
                        proceed = False
                        break
            print('')
            def sendcom(comd):
                out, a, b = '', 0, 5
                shell.send(comd)
                sleep(1)
                print('Waiting for command output')
                while True:
                    a += 1
                    if shell.recv_ready():
                        sleep(1)
                        # print(11)
                        out += shell.recv(10000000).decode()
                        if out.endswith('#') or out.endswith('# '):
                            print(out)
                            return out
                        if out.endswith('[confirm]') or out.endswith('[confirm] '):
                            shell.send('\r')
                            sleep(1)
                            shell.recv(10000)
                            print(out)
                            return out
                        if out.endswith('(yes/no)?') or out.endswith('(yes/no)? '):
                            shell.send('yes')
                            sleep(1)
                            shell.recv(10000)
                            print(out)
                            return out
                        if out.endswith('>') or out.endswith('> '):
                            shell.send('en\rAirtel@123\r')
                            sleep(1)
                            shell.recv(10000)
                            print(out)
                            return out
                    if a > b:
                        print('.',end='.')
                        b += 2
                    sleep(1)
            if proceed==True:
                col = 1
                sheet.max_row
                sheet.cell(row=row, column=col).value = re.split('\t', each)[0]
                log = sendcom(bukconfig)
                sleep(2)
                col += 1
                excellog(row, col, log)
                file.write(log)
                mylog.write(log)
                sendcom(chr(3))


    except (SSHException, socket.error, AuthenticationException):
        col = 1
        sheet.max_row
        sheet.cell(row=row, column=col).value = re.split('\t', each)[0]
        print(re.split('\t', each)[0] + 'SSH Connection failed')
        log = "SSH Connection failed for " + re.split('\t', each)[0]
        col += 1
        excellog(row, col, log)
        file.write(log)
        mylog.write(log)
        wb.save(fType+'/myReport_'+filetime+'.xlsx')
        wb.close()
        connected = False
