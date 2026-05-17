import datetime
import socket
import os
import sys
import re
import paramiko
import string
import xlsxwriter
import openpyxl
from time import sleep
from paramiko import SSHException, AuthenticationException

#password = 'Airtel@123'
password = 'Airtel@1q2w'
username = 'chugorji'
filetime = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
folder = ['/myfiles/', '/FPD/']
home = os.path.dirname(sys.argv[0]).replace('\\', '/')
fpath=''
for ff in folder:
    fpath=home + ff
    if not os.path.exists(fpath):
        os.mkdir(fpath)
#mynodes = open(home + "/myfiles/mynodes").readlines()
mynodes = open(home + "/myfiles/Allnodes").readlines()
myconfig = open(home + "/myfiles/myconfig").readlines()

workbook = xlsxwriter.Workbook(fpath+'MyReport'+filetime+'.xlsx')
worksheet = workbook.add_worksheet('Result')
workbook.close()
wb = openpyxl.load_workbook(fpath+'MyReport'+filetime+'.xlsx')
sheet = wb['Result']


print(mynodes)


def excellog(row, col, log):
    printable = set(string.printable)
    filterlog = ''.join(filter(lambda x: x in printable, log))

    sheet.cell(row=row, column=col).value = filterlog
    wb.save(fpath+'MyReport'+filetime+'.xlsx')
    wb.close()


oneoff = open(fpath + 'oneoff_' + filetime, 'w+')

row = 0
for each in mynodes:
    print(each)
    row += 1
    ip = re.split('\t', each)
    if each.strip() != "" or each.strip() != None or each.strip() != "\n":
        file = open(fpath+ip[0], 'w+')
        cisco = {
            'hostname': ip[1].strip('\n') if len(ip) > 1 else ip[0].strip('\n'),
            'port': 22,
            'username': username,
            'password': password,
            'timeout': 30
        }
    connected = True

    try:
        con = paramiko.SSHClient()
        con.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        con.connect(**cisco)
        waittime = 0
        proceed = True
        with con.invoke_shell() as shell:
            shell.recv(10000)
            print('Wait for connection', '')
            while True:
                if shell.recv_ready():
                    sleep(1)
                    print('')
                    shell.send('terminal len 0\r')
                    print(shell.recv(10000))
                    break
                else:
                    waittime += 1
                    print('.', end="")
                    sleep(1)
                    if waittime >= 25:
                        con.close()
                        proceed = False
                        break
            print('')

            def sendcom(comd):
                out, a, b = '', 0, 5
                shell.send(comd)
                sleep(1)
                print('Waiting for command output')
                while True:
                    a += 1
                    if shell.recv_ready():
                        sleep(1)
                        # print(11)
                        out += shell.recv(1000000).decode()
                        if out.endswith('#') or out.endswith('# ') :
                            print(out)
                            return out
                        if out.endswith('[confirm]') or out.endswith('[confirm] ') or out.endswith('(confirm)') or out.endswith('(confirm) '):
                            shell.send('\r')
                            sleep(1)
                            shell.recv(10000)
                            print(out)
                            return out
                        if out.endswith('(yes/no)?') or out.endswith('(yes/no)? ') or out.endswith('])?') or out.endswith('])? '):
                            shell.send('yes')
                            sleep(1)
                            shell.recv(10000)
                            print(out)
                            return out
                        if out.endswith('>') or out.endswith('> ') :
                            shell.send('en\rAirtel@123\r')
                            sleep(1)
                            shell.recv(10000)
                            print(out)
                            return out

                        if out.endswith('ord:') or out.endswith('ord: ') :
                            shell.send('Airtel@1q2w\r')
                            sleep(1)
                            shell.recv(10000)
                            print(out)
                            return out

                    if a > b:
                        print('.', end='.')
                        b += 2
                    sleep(1)

            col = 1
            sheet.max_row
            sheet.cell(row=row, column=col).value = ip[0] 
            perNodelog = ''
            for each in myconfig:
                if each != None and proceed==True :
                    log = sendcom(each)
                    sleep(2)
                    col += 1
                    excellog(row, col, log)
                    file.write(log)
                    perNodelog += log
            oneoff.write('\n\n===============\n' +
                         ' '.join(ip) + '\n===============\n' + perNodelog)
            sendcom(chr(3))

    except (SSHException, socket.error, AuthenticationException):
        col = 1
        sheet.max_row
        sheet.cell(row=row, column=col).value = re.split('\t', each)[0]
        print(re.split('\t', each)[0] + 'SSH Connection failed')
        log = "SSH Connection failed for " + re.split('\t', each)[0]
        col += 1
        excellog(row, col, log)
        file.write(log)
        wb.save(fpath+'MyReport'+filetime+'.xlsx')
        wb.close()
        connected = False
